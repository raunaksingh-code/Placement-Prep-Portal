"""Extract JD PDFs and the interview question tracker into structured JSON.

Run once (or after the source files change):
    python -m app.seed.extract_jd

Writes data/seed_source/jd_extract/job_descriptions.json and question_bank.json.
The seed script reads those files; it never touches the PDFs or the workbook.
"""

import json
import re
import unicodedata
from pathlib import Path

import fitz
import openpyxl

SEED_DIR = Path(__file__).resolve().parents[3] / "data" / "seed_source"
JD_DIR = SEED_DIR / "JDs"
TRACKER = SEED_DIR / "DM2-Final Placements'2026 - Question Tracker.xlsx"
OUT_DIR = SEED_DIR / "jd_extract"

# Folders whose files are named "Company_Role.pdf" rather than grouping by company folder
FLAT_FOLDERS = {"Single JD_Company Wise"}

# The JD folders and the tracker spell some companies differently. Map both spellings
# to one canonical name so a company page shows its JDs and its questions together.
# Deloitte India and Deloitte USI are deliberately NOT merged - they recruit separately.
COMPANY_ALIASES = {
    "airtel africa llp": "Airtel Africa",
    "alvarez & marsal gcc": "Alvarez & Marsal",
    "hcl technologies": "HCL Tech",
    "quickclean laundry systems": "Quick Clean Laundry Systems",
    "bofa": "Bank of America",
    "newgen software": "Newgen",
}


def canonical_company(name: str) -> str:
    return COMPANY_ALIASES.get(name.strip().lower(), name.strip())

# Headings that mark a new section in a JD. Order matters only for matching, not output.
SECTION_PATTERNS = [
    ("responsibilities", r"(key\s+)?(job\s+)?responsibilit|role\s*&?\s*responsibilit|what\s+you.{0,5}ll\s+do|job\s+purpose|about\s+the\s+role|role\s+description|job\s+description"),
    ("qualifications", r"qualification|eligibilit|education|who\s+we.{0,5}re\s+looking|candidate\s+profile|desired\s+profile|requirement"),
    ("skills", r"skill|competenc|attribute|proficienc"),
    ("experience", r"experience|work\s+ex"),
    ("about", r"about\s+(us|the\s+company|the\s+team)|company\s+(overview|profile)"),
    ("compensation", r"compensation|salary|ctc|package|benefit"),
    ("location", r"location|work\s+location|base\s+location"),
]

SKILL_KEYWORDS = [
    "SQL", "Python", "R", "Excel", "Power BI", "Tableau", "SAP", "VBA", "Alteryx",
    "Machine Learning", "Advanced Excel", "MS Office", "PowerPoint", "Looker",
    "Salesforce", "CRM", "ERP", "Jira", "Agile", "Scrum", "AWS", "Azure",
    "Financial Modelling", "Financial Modeling", "Valuation", "Forecasting",
    "Data Analysis", "Data Visualization", "Data Visualisation", "Business Analysis",
    "Stakeholder Management", "Project Management", "Presales", "Market Research",
    "Communication", "Negotiation", "Presentation", "Problem Solving",
    "Client Management", "Consulting", "Analytics", "Digital Marketing", "SEO",
]

# Role keyword -> category. First match wins, so order from most to least specific.
ROLE_CATEGORIES = [
    ("Consulting", r"consultant|consulting|strategy|transformation|advisory"),
    ("Data & Analytics", r"analyst|analytics|data|research|insight|bi\b|business intelligence"),
    ("Finance", r"financ|invest|wealth|banking|treasury|audit|risk|fp&a|accounts|credit"),
    ("Sales & Business Development", r"sales|business development|bd\b|account manager|relationship manager|client success|customer success"),
    ("Marketing", r"marketing|brand|growth|communications|content"),
    ("Product & Technology", r"product|technology|technical|engineer|developer|platform|sap|it\b"),
    ("Operations", r"operation|supply chain|logistics|process|delivery|pmo|program manager|project manag"),
    ("Human Resources", r"\bhr\b|human resource|talent|recruit|people"),
    ("General Management", r"management trainee|\bmt\b|general management|graduate trainee|leadership"),
]

from app.seed.category_map import CATEGORY_TOPICS  # noqa: E402


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^\w\s-]", " ", text).strip().lower()
    return re.sub(r"[\s_-]+", "-", text)


def clean_text(raw: str) -> str:
    """Normalise PDF text: fix bullets, collapse the line breaks PDFs scatter mid-sentence."""
    text = raw.replace("•", "\n• ").replace("", "\n• ")
    text = unicodedata.normalize("NFKD", text)
    text = text.replace("’", "'").replace("‘", "'")
    text = text.replace("“", '"').replace("”", '"')
    text = text.replace("–", "-").replace("—", "-")
    text = "".join(c for c in text if c.isprintable() or c == "\n")

    lines = [ln.strip() for ln in text.split("\n")]
    out: list[str] = []
    for ln in lines:
        if not ln:
            out.append("")
            continue
        # A lone bullet glyph belongs with the text on the following line
        if ln == "•":
            out.append("•")
            continue
        if out and out[-1] == "•":
            out[-1] = f"• {ln}"
            continue
        # Continuation of a wrapped sentence: previous line has no terminator
        if (
            out
            and out[-1]
            and not out[-1].endswith((".", ":", "?", "!"))
            and not ln.startswith("•")
            and not out[-1].startswith("#")
            and len(out[-1]) > 40
        ):
            out[-1] = f"{out[-1]} {ln}"
        else:
            out.append(ln)
    collapsed = "\n".join(out)
    return re.sub(r"\n{3,}", "\n\n", collapsed).strip()


def is_heading(line: str) -> str | None:
    """Return the section key if the line looks like a section heading."""
    stripped = line.strip().rstrip(":").strip()
    if not stripped or len(stripped) > 70 or stripped.startswith("•"):
        return None
    # Headings are short and rarely end in a full stop
    if stripped.endswith("."):
        return None
    low = stripped.lower()
    for key, pattern in SECTION_PATTERNS:
        if re.search(pattern, low):
            # Guard against a long sentence that merely mentions the word
            if len(stripped.split()) <= 8:
                return key
    return None


def split_sections(text: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current = "overview"
    sections[current] = []
    for line in text.split("\n"):
        key = is_heading(line)
        if key:
            current = key
            sections.setdefault(current, [])
            continue
        sections.setdefault(current, []).append(line)
    return {
        k: re.sub(r"\n{3,}", "\n\n", "\n".join(v).strip())
        for k, v in sections.items()
        if "".join(v).strip()
    }


def detect_skills(text: str) -> list[str]:
    low = text.lower()
    found = []
    for skill in SKILL_KEYWORDS:
        if re.search(rf"\b{re.escape(skill.lower())}\b", low):
            # Normalise the two spellings we accept
            canonical = skill.replace("Modeling", "Modelling").replace("Visualization", "Visualisation")
            if canonical not in found:
                found.append(canonical)
    return found


def categorise(role: str, text: str) -> str:
    """Categorise by role title first; the JD body is only a fallback.

    Body text mentions words like "strategy" constantly, so matching it with equal
    weight pulls almost everything into Consulting.
    """
    role_low = role.lower()
    for category, pattern in ROLE_CATEGORIES:
        if re.search(pattern, role_low):
            return category
    body = text[:1200].lower()
    for category, pattern in ROLE_CATEGORIES:
        if re.search(pattern, body):
            return category
    return "General Management"


def tidy_role(name: str) -> str:
    """Turn a filename stem into a readable role title."""
    role = name.replace("_", " - ").strip()
    role = re.sub(r"\s*-\s*", " - ", role)
    role = re.sub(r"\s{2,}", " ", role)
    return role.strip(" -")


def parse_jd_file(path: Path) -> dict | None:
    rel = path.relative_to(JD_DIR)
    parts = list(rel.parts)
    stem = path.stem

    if parts[0] in FLAT_FOLDERS:
        # Files are "Company_Role.pdf", but some use " -" as the separator instead
        # ("Repos Energy -Area Sales Lead.pdf"). Split on whichever comes first.
        underscore = stem.find("_")
        dash = stem.find(" -")
        candidates = [i for i in (underscore, dash) if i != -1]
        if candidates:
            cut = min(candidates)
            company_raw = stem[:cut]
            role_raw = stem[cut:].lstrip("_").lstrip(" -").strip()
        else:
            company_raw, role_raw = stem, "Role"
    else:
        company_raw = parts[0]
        # Nested "Deloitte India 1" style folders repeat the company name; drop the index
        if "_" in stem and len(parts) == 2:
            maybe_company, maybe_role = stem.split("_", 1)
            if slugify(maybe_company).startswith(slugify(company_raw)[:8]):
                role_raw = maybe_role
            else:
                role_raw = stem
        else:
            role_raw = stem

    # Drop file-index suffixes the source uses to number multiple JDs: "HCL 2", "Z1 Tech 1 & 2"
    company_raw = re.sub(r"\s+\d+(\s*&\s*\d+)*$", "", company_raw).strip()
    company = canonical_company(company_raw)
    role = tidy_role(role_raw)

    doc = fitz.open(path)
    raw = "\n".join(page.get_text() for page in doc)
    doc.close()
    text = clean_text(raw)
    if len(text) < 120:
        return None

    sections = split_sections(text)
    return {
        "company": company,
        "company_slug": slugify(company),
        "role": role,
        "slug": f"{slugify(company)}--{slugify(role)}",
        "category": categorise(role, text),
        "skills": detect_skills(text),
        "sections": sections,
        "full_text": text,
        "source_file": str(rel).replace("\\", "/"),
    }


def extract_jds() -> list[dict]:
    jds: list[dict] = []
    seen: set[str] = set()
    for path in sorted(JD_DIR.rglob("*.pdf")):
        try:
            jd = parse_jd_file(path)
        except Exception as exc:  # noqa: BLE001 - report and continue
            print(f"  ! failed {path.name}: {exc}")
            continue
        if not jd:
            print(f"  ! too little text, skipped: {path.name}")
            continue
        slug = jd["slug"]
        n = 2
        while slug in seen:
            slug = f"{jd['slug']}-{n}"
            n += 1
        jd["slug"] = slug
        seen.add(slug)
        jds.append(jd)
    return jds


ROUND_COLUMNS = [
    (3, "L1 Interview"),
    (4, "L2 Interview"),
    (5, "L3 Interview"),
    (6, "HR Round"),
    (7, "Leadership Round"),
]

QUESTION_CATEGORIES = [
    ("Guesstimate", r"guesstimate|estimate the number|estimate how many|market siz"),
    (
        "Aptitude & Logic",
        r"logical reasoning|puzzle|probability|permutation|milk-and-water|aptitude"
        r"|brain\s*teaser|riddle|sequence|series problem|solve this problem",
    ),
    ("Case Study", r"\bcase\b|case study|case interview|situational question"),
    (
        "Resume & Experience",
        r"tell me about yourself|tmay|walk me through|your internship|your resume"
        r"|your project|work experience|your ppt|your presentation|best ppt"
        r"|your role at|your academic|about your (education|background|college)",
    ),
    (
        "Technical",
        r"sql|excel|python|\bvba\b|power\s*bi|tableau|vlookup|pivot|macro|dashboard"
        r"|formula|financial statement|valuation|\bdcf\b|\birr\b|\bmoic\b|\bnpv\b"
        r"|accounting|balance sheet|cash flow|bond|equity|derivative|hedge"
        r"|lean|six sigma|\b5s\b|iso \d|blockchain|\bapi\b|machine learning"
        r"|working capital|depreciation|amortis|ebitda|p&l|debit|credit note"
        r"|cross-?sell vs upsell|waterfall|clawback|zero-coupon"
        r"|regression|statistic|\bp2p\b|\bo2c\b|\br2r\b|coding|algorithm"
        r"|risk assessment|risk management|\bgst\b|taxation|supply chain",
    ),
    (
        "HR & Behavioural",
        r"why do you want|why this|why join|why should we|why are you|why did you"
        r"|why \w+\?|why \w+ /|strength|weakness|relocat|salary|expectation"
        r"|conflict|work in a team|leadership|challenge|failure|greatest achievement"
        r"|(five|5) years|long[- ]term goal|career goal|see yourself"
        r"|comfortable with (travel|shift|night)|notice period|other offers",
    ),
    (
        "Personal & Background",
        r"family background|where are you from|hometown|favourite city|which city"
        r"|cat percentile|your (10th|12th|graduation) (marks|percentage|score)"
        r"|low percentage|the \d+-year gap|your hobbies|what do you do in your free",
    ),
    (
        "Domain & Industry",
        r"industry|market trend|sector|competitor|business model|company do"
        r"|about the company|about (us|our)|\bceo\b|verticals|our products"
        r"|current affairs|recent news|green financ|\besg\b|what do you know about"
        r"|difference between a? ?(current|savings)|our clients|our services",
    ),
    ("Group Discussion", r"\bgd\b|group discussion|gd topic|debate topic"),
    (
        "Situational & Judgement",
        r"^situation|situation:|how would you handle|how do you handle|what would you do if"
        r"|your subordinate|your team member|a client (asks|refuses|is angry)|deadline",
    ),
    (
        "Role & Motivation",
        r"why (marketing|finance|sales|consulting|analytics|operations|hr)"
        r"|this role|the role|inclined more towards|prefer .* or |which (tool|team|domain)"
        r"|first 90 days|align with this|excite you|interested in this",
    ),
]


def categorise_question(text: str) -> str:
    low = text.lower()
    for category, pattern in QUESTION_CATEGORIES:
        if re.search(pattern, low):
            return category
    return "General"


def split_questions(cell: str) -> list[dict]:
    """Split one round's cell into individual questions.

    Lines begin with • or ⭐ (starred = frequently reported). The first line is
    usually a header like "L1 Interview (R1):" and notes blocks are kept as context.
    """
    if not cell:
        return []
    text = unicodedata.normalize("NFKD", str(cell))
    text = text.replace("’", "'").replace("‘", "'")
    text = text.replace("“", '"').replace("”", '"')
    text = text.replace("–", "-").replace("—", "-")

    items: list[dict] = []
    is_note_block = False
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        starred = "⭐" in line
        line = line.replace("⭐", "").strip()
        # Header lines: "L1 Interview (R1):", "Presentation (R2):", "Round 2:", "Notes:"
        if line.endswith(":") and len(line) < 60:
            if re.match(r"^notes\b", line, re.I):
                is_note_block = True
                continue
            if re.search(r"\(r\d\)|^(l[123]\b|round\s*\d|interview\s*\d)", line, re.I) or re.match(
                r"^(l[123]\s*interview|hr\s*round|leadership|personal interview|presentation"
                r"|group discussion|gd|case study round|technical round|assessment)\b",
                line,
                re.I,
            ):
                is_note_block = False
                continue
        line = line.lstrip("•").strip()
        line = re.sub(r"^[•\-•\s]+", "", line).strip()
        if len(line) < 8:
            continue
        items.append(
            {
                "text": line,
                "starred": starred,
                "is_note": is_note_block,
                "category": categorise_question(line),
            }
        )
    return items


def extract_question_bank() -> list[dict]:
    wb = openpyxl.load_workbook(TRACKER, data_only=True)
    entries: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            company = canonical_company(
                (row[1] or sheet_name or "").replace("\xa0", " ").strip()
            )
            role = (row[2] or "").replace("\xa0", " ").strip()
            if not company:
                continue
            rounds = []
            for idx, round_name in ROUND_COLUMNS:
                if idx >= len(row):
                    continue
                questions = split_questions(row[idx])
                if questions:
                    rounds.append({"round": round_name, "questions": questions})
            if not rounds:
                continue
            entries.append(
                {
                    "company": company,
                    "company_slug": slugify(company),
                    "role": role or "General",
                    "category": categorise(role or company, role or ""),
                    "rounds": rounds,
                }
            )
    return entries


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Extracting job descriptions...")
    jds = extract_jds()
    (OUT_DIR / "job_descriptions.json").write_text(
        json.dumps(jds, indent=1, ensure_ascii=False), encoding="utf-8"
    )
    companies = {j["company_slug"] for j in jds}
    print(f"  {len(jds)} JDs across {len(companies)} companies")

    print("Extracting interview question bank...")
    bank = extract_question_bank()
    (OUT_DIR / "question_bank.json").write_text(
        json.dumps(bank, indent=1, ensure_ascii=False), encoding="utf-8"
    )
    total_q = sum(len(r["questions"]) for e in bank for r in e["rounds"])
    print(f"  {total_q} questions across {len(bank)} company-role entries")

    (OUT_DIR / "category_topics.json").write_text(
        json.dumps(CATEGORY_TOPICS, indent=1), encoding="utf-8"
    )
    print("Wrote", OUT_DIR)


if __name__ == "__main__":
    main()
